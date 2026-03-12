#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sqlite3
import subprocess
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ACCESS_TABLES = [
    ("WELL_SEARCH_INDEX", "well_search_index", "wa_num"),
    ("PRD_Profile_Gas", "prd_profile_gas", "wa_num"),
    ("ADV_Profile_Gas", "adv_profile_gas", "wa_num"),
    ("FRAC_Summary", "frac_summary", "wa_num"),
    ("DIR_SURVEY", "dir_survey", "wa_num"),
    ("DRILL_EVENTS", "drill_events", "wa_num"),
    ("FRAC_Descriptions", "frac_descriptions", "wa_num"),
    ("Gas_Analysis", "gas_analysis", "wa_num"),
    ("ABANDON", "abandon", "wa_num"),
    ("CASING", "casing", "wa_num"),
    ("PAY_ZONE", "pay_zone", "wa_num"),
]

WORKBOOK_OVERLAYS = [
    ("PRD_Profile_Gas", "prd_profile_gas", "wa_num", {}),
    ("ADV_Profile_Gas", "adv_profile_gas", "wa_num", {}),
    ("FRAC_Summary", "frac_summary", "wa_num", {}),
    ("DIR_SURVEY", "dir_survey", "wa_num", {}),
    ("Drilling_Events", "drill_events", "wa_num", {}),
    ("FRAC_Descriptions", "frac_descriptions", "wa_num", {}),
    ("GAS_ANAL", "gas_analysis", "wa_num", {}),
    ("ABANDON", "abandon", "wa_num", {"order": "position"}),
    ("CASINGS", "casing", "wa_num", {"order": "position"}),
    ("PAY_ZONE", "pay_zone", "wa_num", {}),
]

NULL_SENTINEL = "__MDB_NULL__"
HEADER_CLEAN_RE = re.compile(r"[^a-zA-Z0-9]+")


@dataclass
class SheetSchema:
    column_names: list[str]
    source_indexes: list[int]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import BCER source files into SQLite.")
    parser.add_argument("--xlsx", required=True, help="Path to the source .xlsm workbook.")
    parser.add_argument("--db", required=True, help="Path to the destination SQLite database.")
    parser.add_argument("--accdb", required=True, help="Path to the source .accdb database.")
    return parser.parse_args()


def normalize_identifier(raw: str) -> str:
    candidate = HEADER_CLEAN_RE.sub("_", raw.strip()).strip("_").lower()
    return candidate or "column"


def build_sheet_schema(
    header_row: Iterable[object],
    *,
    aliases: dict[str, str] | None = None,
    ignored_columns: set[str] | None = None,
) -> SheetSchema:
    aliases = aliases or {}
    ignored_columns = ignored_columns or set()
    seen: set[str] = set()
    column_names: list[str] = []
    source_indexes: list[int] = []

    for index, header in enumerate(header_row):
        if header in (None, ""):
            continue

        normalized = normalize_identifier(str(header))
        normalized = aliases.get(normalized, normalized)
        if normalized in ignored_columns:
            continue

        unique = normalized
        suffix = 2
        while unique in seen:
            unique = f"{normalized}_{suffix}"
            suffix += 1

        seen.add(unique)
        column_names.append(unique)
        source_indexes.append(index)

    return SheetSchema(column_names=column_names, source_indexes=source_indexes)


def quote_identifier(identifier: str) -> str:
    return f'"{identifier}"'


def create_table(cursor: sqlite3.Cursor, table_name: str, schema: SheetSchema) -> None:
    column_sql = ", ".join(f'{quote_identifier(name)} NUMERIC' for name in schema.column_names)
    cursor.execute(f"DROP TABLE IF EXISTS {quote_identifier(table_name)}")
    cursor.execute(f"CREATE TABLE {quote_identifier(table_name)} ({column_sql})")


def insert_batch(connection: sqlite3.Connection, table_name: str, schema: SheetSchema, batch: list[tuple[object, ...]]) -> None:
    if not batch:
        return

    placeholders = ", ".join("?" for _ in schema.column_names)
    columns_sql = ", ".join(quote_identifier(name) for name in schema.column_names)
    sql = f"INSERT INTO {quote_identifier(table_name)} ({columns_sql}) VALUES ({placeholders})"
    connection.executemany(sql, batch)


def create_index(cursor: sqlite3.Cursor, table_name: str, column_name: str, unique: bool = False) -> None:
    index_name = f"idx_{table_name}_{column_name}"
    uniqueness = "UNIQUE " if unique else ""
    cursor.execute(
        f"CREATE {uniqueness}INDEX IF NOT EXISTS {quote_identifier(index_name)} "
        f"ON {quote_identifier(table_name)} ({quote_identifier(column_name)})"
    )


def delete_wa_rows(connection: sqlite3.Connection, table_name: str, wa_column: str, rows: list[tuple[object, ...]], schema: SheetSchema) -> None:
    if wa_column not in schema.column_names:
        return

    wa_index = schema.column_names.index(wa_column)
    wa_values = sorted({row[wa_index] for row in rows if row[wa_index] not in (None, "")})
    if not wa_values:
        return

    placeholders = ", ".join("?" for _ in wa_values)
    connection.execute(
        f"DELETE FROM {quote_identifier(table_name)} WHERE {quote_identifier(wa_column)} IN ({placeholders})",
        wa_values,
    )


def store_metadata(cursor: sqlite3.Cursor, source_row: dict[str, object], about_paragraphs: list[str]) -> None:
    cursor.execute("DROP TABLE IF EXISTS metadata")
    cursor.execute("CREATE TABLE metadata (key TEXT PRIMARY KEY, value TEXT NOT NULL)")

    metadata_items = {
        "author_name": source_row.get("author_name", "") or "",
        "author_email": source_row.get("author_email", "") or "",
        "source_agency": source_row.get("source_agency", "") or "",
        "source_website": source_row.get("source_website", "") or "",
        "data_current_to": source_row.get("data_current_to", "") or "",
        "import_timestamp": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "about_paragraphs": json.dumps(about_paragraphs),
    }

    cursor.executemany(
        "INSERT INTO metadata (key, value) VALUES (?, ?)",
        list(metadata_items.items()),
    )


def store_dataset_counts(cursor: sqlite3.Cursor, counts: dict[str, int]) -> None:
    cursor.execute("DROP TABLE IF EXISTS dataset_counts")
    cursor.execute("CREATE TABLE dataset_counts (table_name TEXT PRIMARY KEY, row_count INTEGER NOT NULL)")
    cursor.executemany(
        "INSERT INTO dataset_counts (table_name, row_count) VALUES (?, ?)",
        list(counts.items()),
    )


def collect_about_paragraphs(workbook) -> list[str]:
    worksheet = workbook["A_Word"]
    paragraphs: list[str] = []

    for row in worksheet.iter_rows(values_only=True):
        values = [str(value).strip() for value in row if value not in (None, "")]
        if values:
            paragraphs.append(" ".join(values))

    return paragraphs


def read_source_row(workbook) -> dict[str, object]:
    worksheet = workbook["Source"]
    schema = build_sheet_schema(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    values = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))
    return {
        column_name: values[source_index]
        for column_name, source_index in zip(schema.column_names, schema.source_indexes, strict=True)
    }


def read_code_lookup_rows(worksheet, *, code_column: int, description_column: int, start_row: int) -> list[tuple[object, object]]:
    rows: list[tuple[object, object]] = []
    current_row = start_row

    while current_row <= worksheet.max_row:
        code = worksheet.cell(current_row, code_column).value
        description = worksheet.cell(current_row, description_column).value
        if code in (None, "") and description in (None, ""):
            break

        rows.append((code, description))
        current_row += 1

    return rows


def import_code_lookups(connection: sqlite3.Connection, workbook) -> None:
    worksheet = workbook["CODES"]
    cursor = connection.cursor()

    for table_name, code_column, description_column in [
        ("area_codes", 1, 2),
        ("formation_codes", 4, 5),
    ]:
        rows = read_code_lookup_rows(
            worksheet,
            code_column=code_column,
            description_column=description_column,
            start_row=3,
        )
        cursor.execute(f"DROP TABLE IF EXISTS {quote_identifier(table_name)}")
        cursor.execute(
            f"CREATE TABLE {quote_identifier(table_name)} "
            '(code NUMERIC NOT NULL, description TEXT)'
        )
        connection.executemany(
            f"INSERT INTO {quote_identifier(table_name)} (code, description) VALUES (?, ?)",
            rows,
        )
        create_index(cursor, table_name, "code", unique=True)


def import_workbook_table(
    connection: sqlite3.Connection,
    workbook,
    sheet_name: str,
    table_name: str,
    wa_column: str | None,
    *,
    aliases: dict[str, str] | None = None,
    create: bool = True,
) -> int:
    cursor = connection.cursor()
    worksheet = workbook[sheet_name]
    schema = build_sheet_schema(
        next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)),
        aliases=aliases,
    )

    if create:
        create_table(cursor, table_name, schema)

    rows: list[tuple[object, ...]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        values = tuple(row[index] for index in schema.source_indexes)
        if all(value in (None, "") for value in values):
            continue
        rows.append(values)

    if not rows:
        return 0

    if not create and wa_column:
        delete_wa_rows(connection, table_name, wa_column, rows, schema)

    insert_batch(connection, table_name, schema, rows)

    if wa_column and wa_column in schema.column_names and create:
        create_index(cursor, table_name, wa_column)

    return len(rows)


def import_access_table(connection: sqlite3.Connection, accdb_path: Path, source_table: str, table_name: str, wa_column: str | None) -> int:
    if shutil.which("mdb-export") is None:
        raise RuntimeError("mdb-export is required to import .accdb data. Install mdbtools first.")

    process = subprocess.Popen(
        [
            "mdb-export",
            "-e",
            "-0",
            NULL_SENTINEL,
            str(accdb_path),
            source_table,
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
    )

    if process.stdout is None:
        raise RuntimeError(f"Unable to read rows from {source_table}.")

    reader = csv.reader(process.stdout)

    try:
        header_row = next(reader)
    except StopIteration as exc:
        stderr = process.stderr.read() if process.stderr else ""
        raise RuntimeError(f"No rows returned for {source_table}. {stderr}") from exc

    schema = build_sheet_schema(header_row, ignored_columns={"id"})
    cursor = connection.cursor()
    create_table(cursor, table_name, schema)

    batch: list[tuple[object, ...]] = []
    row_count = 0

    for raw_row in reader:
        values = []
        for index in schema.source_indexes:
            value = raw_row[index] if index < len(raw_row) else None
            if value == NULL_SENTINEL or value == "":
                values.append(None)
            else:
                values.append(value)

        batch.append(tuple(values))

        if len(batch) == 1000:
            insert_batch(connection, table_name, schema, batch)
            row_count += len(batch)
            batch.clear()

    insert_batch(connection, table_name, schema, batch)
    row_count += len(batch)

    process.stdout.close()
    stderr = process.stderr.read() if process.stderr else ""
    return_code = process.wait()
    if return_code != 0:
        raise RuntimeError(f"mdb-export failed for {source_table}: {stderr.strip()}")

    if wa_column and wa_column in schema.column_names:
        create_index(cursor, table_name, wa_column)

    return row_count


def query_table_counts(connection: sqlite3.Connection, table_names: list[str]) -> dict[str, int]:
    counts: dict[str, int] = {}
    cursor = connection.cursor()
    for table_name in table_names:
        counts[table_name] = cursor.execute(f"SELECT COUNT(*) FROM {quote_identifier(table_name)}").fetchone()[0]
    return counts


def build_well_search(cursor: sqlite3.Cursor) -> None:
    prod_3yr = " + ".join(f"COALESCE(mprd_{index:03d}, 0)" for index in range(1, 37))
    prod_5yr = " + ".join(f"COALESCE(mprd_{index:03d}, 0)" for index in range(1, 61))

    cursor.execute("DROP TABLE IF EXISTS well_search")
    cursor.execute(
        f"""
        CREATE TABLE well_search AS
        WITH prod AS (
          SELECT
            CAST(wa_num AS INTEGER) AS wa_num,
            ({prod_3yr}) AS gas_prod_3yr,
            ({prod_5yr}) AS gas_prod_5yr,
            CAST(first_prod_period AS INTEGER) AS first_prod_period
          FROM prd_profile_gas
        ),
        uwi AS (
          SELECT
            CAST(wa_num AS INTEGER) AS wa_num,
            GROUP_CONCAT(DISTINCT uwi) AS uwi_list
          FROM adv_profile_gas
          WHERE uwi IS NOT NULL AND TRIM(CAST(uwi AS TEXT)) <> ''
          GROUP BY CAST(wa_num AS INTEGER)
        )
        SELECT
          CAST(wsi.wa_num AS INTEGER) AS wa_num,
          wsi.well_name AS well_name,
          wsi.operator AS operator,
          CAST(wsi.oper_id AS INTEGER) AS operator_id,
          wsi.oper_abrv AS operator_abbr,
          uwi.uwi_list AS uwi_list,
          CAST(wsi.area_num AS INTEGER) AS area_code,
          wsi.area_txt AS area_desc,
          CAST(wsi.form_num AS INTEGER) AS form_code,
          wsi.form_txt AS form_desc,
          CAST(wsi.spud_mon AS INTEGER) AS spud_mon,
          CAST(wsi.rig_rel_mon AS INTEGER) AS rig_rel_mon,
          CAST(wsi.first_prod_mon AS INTEGER) AS first_prod_mon,
          NULLIF(wsi.horiz_well, '-') AS orientation,
          CAST(wsi.surf_lat_dec_deg AS REAL) AS surf_lat,
          CAST(wsi.surf_lon_dec_deg AS REAL) AS surf_lon,
          NULL AS grid,
          COALESCE(prod.gas_prod_3yr, 0) AS gas_prod_3yr,
          COALESCE(prod.gas_prod_5yr, 0) AS gas_prod_5yr,
          prod.first_prod_period AS first_prod_period
        FROM well_search_index wsi
        LEFT JOIN prod ON prod.wa_num = CAST(wsi.wa_num AS INTEGER)
        LEFT JOIN uwi ON uwi.wa_num = CAST(wsi.wa_num AS INTEGER)
        WHERE wsi.wa_num IS NOT NULL
        """
    )

    create_index(cursor, "well_search", "wa_num", unique=True)
    create_index(cursor, "well_search", "well_name")
    create_index(cursor, "well_search", "operator")
    create_index(cursor, "well_search", "operator_id")
    create_index(cursor, "well_search", "area_desc")
    create_index(cursor, "well_search", "form_desc")
    create_index(cursor, "well_search", "spud_mon")
    create_index(cursor, "well_search", "rig_rel_mon")
    create_index(cursor, "well_search", "first_prod_mon")
    create_index(cursor, "well_search", "gas_prod_3yr")
    create_index(cursor, "well_search", "gas_prod_5yr")
    create_index(cursor, "well_search", "orientation")
    create_index(cursor, "well_search", "surf_lat")
    create_index(cursor, "well_search", "surf_lon")


def import_sources(xlsx_path: Path, db_path: Path, accdb_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True, keep_vba=False)
    connection = sqlite3.connect(db_path)

    try:
        connection.execute("PRAGMA journal_mode = WAL")
        connection.execute("PRAGMA synchronous = NORMAL")
        cursor = connection.cursor()

        for source_table, table_name, wa_column in ACCESS_TABLES:
            import_access_table(connection, accdb_path, source_table, table_name, wa_column)

        for sheet_name, table_name, wa_column, aliases in WORKBOOK_OVERLAYS:
            import_workbook_table(
                connection,
                workbook,
                sheet_name,
                table_name,
                wa_column,
                aliases=aliases,
                create=False,
            )

        import_workbook_table(connection, workbook, "CODES", "codes", None)
        import_code_lookups(connection, workbook)

        source_row = read_source_row(workbook)
        about_paragraphs = collect_about_paragraphs(workbook)
        store_metadata(cursor, source_row, about_paragraphs)

        build_well_search(cursor)

        dataset_counts = query_table_counts(
            connection,
            [
                "abandon",
                "adv_profile_gas",
                "casing",
                "codes",
                "area_codes",
                "dir_survey",
                "drill_events",
                "formation_codes",
                "frac_descriptions",
                "frac_summary",
                "gas_analysis",
                "pay_zone",
                "prd_profile_gas",
                "well_search_index",
                "well_search",
            ],
        )
        store_dataset_counts(cursor, dataset_counts)
        connection.commit()
    finally:
        workbook.close()
        connection.close()


def main() -> None:
    args = parse_args()
    import_sources(
        Path(args.xlsx).resolve(),
        Path(args.db).resolve(),
        Path(args.accdb).resolve(),
    )


if __name__ == "__main__":
    main()
